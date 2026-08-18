
import pandas as pd
import numpy as np
from typing import List, Dict, Set, Any
from copy import deepcopy

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import openpyxl
import os
# ===============================
# CONFIGURE THESE
# ===============================

def getting_endpoint_chunk_flow(INPUT_XLSX,OUTPUT_PATH,OUTPUT_XLSX,INPUT_SHEET):
    OUTPUT_XLSX = os.path.join(OUTPUT_PATH,OUTPUT_XLSX)
                               # sheet index or name, e.g., "Sheet1"
    INCLUDE_LAST_GROUPS_ONLY_LEVEL = True     # if a node has no children but has groups, emit groups-only level
    MAX_DEPTH = 50                            # hard stop to avoid pathological cycles

    # ✅ Merge ALL parent levels (root_chunk, Level_1, Level_2, ...)
    MERGE_PARENT_CELLS_MODE = "all_parents"   # "root_only" or "all_parents"

    # ===============================
    # 1) Load data from Excel
    # ===============================
    df = pd.read_excel(INPUT_XLSX, sheet_name=INPUT_SHEET, engine="openpyxl")

    # Ensure required columns exist (create empty if missing)
    for col in ["chunk_id", "direct_children", "groups"]:
        if col not in df.columns:
            df[col] = ""

    # ===============================
    # 2) Helpers to parse list-like cells
    # ===============================
    def _is_missing(x) -> bool:
        if x is None:
            return True
        if isinstance(x, float) and np.isnan(x):
            return True
        s = str(x).strip()
        return s == "" or s.lower() in {"nan", "nil", "none", "null", "0"}

    def _split_tokens(x: Any) -> List[str]:
        """
        Split by common separators and clean tokens.
        Handles 'C001, C002', 'C001|C002', 'C001\\nC002', etc.
        Ignores empty, 'Nil', 'NaN', '0', etc.
        """
        if _is_missing(x):
            return []
        s = str(x).strip()
        for sep in ["|", ";", "\n", "\t"]:
            s = s.replace(sep, ",")
        raw_tokens = [t.strip() for t in s.split(",")]
        tokens = []
        for t in raw_tokens:
            if not _is_missing(t):
                tokens.append(t)
        return tokens

    def parse_children_cell(x) -> List[str]:
        return _split_tokens(x)

    def parse_groups_cell(x) -> List[str]:
        return _split_tokens(x)

    def dedup_preserve(seq: List[str]) -> List[str]:
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    # ===============================
    # 3) Build adjacency and group maps
    # ===============================
    def build_graph(source_df: pd.DataFrame) -> tuple[Dict[str, List[str]], Dict[str, List[str]]]:
        adj_map: Dict[str, List[str]] = {}
        group_map: Dict[str, List[str]] = {}

        for _, row in source_df.iterrows():
            cid = str(row["chunk_id"]).strip()
            children = dedup_preserve(parse_children_cell(row.get("direct_children", "")))
            groups = dedup_preserve(parse_groups_cell(row.get("groups", "")))
            adj_map[cid] = children
            group_map[cid] = groups

        return adj_map, group_map

    # ===============================
    # 4) Determine roots (Level 1)
    # ===============================
    def find_roots(adj_map: Dict[str, List[str]]) -> List[str]:
        all_nodes = set(adj_map.keys())
        all_children = set()
        for kids in adj_map.values():
            all_children.update(kids)
        roots = sorted(all_nodes - all_children)
        return roots

    # ===============================
    # 5) Row-wise expanding flow (explode children to rows, de-dup per parent)
    # ===============================
    def build_levelwise_flow_rows(source_df: pd.DataFrame,
                                include_last_groups_only_level: bool = True,
                                max_depth: int = 50) -> pd.DataFrame:
        adj_map, group_map = build_graph(source_df)
        roots = find_roots(adj_map)

        records = []

        # Row state carries output row dict + traversal frontier + visited
        class RowState:
            def __init__(self, root: str):
                self.out = {"root_chunk": root, "Level_1": root}
                self.frontier = [root]       # nodes of the previous level for this row
                self.visited = set([root])   # prevent cycles

        for root in roots:
            # start with one row for the root
            current_rows = [RowState(root)]
            level_idx = 2
            depth_guard = 0

            while current_rows and depth_guard < max_depth:
                next_rows = []
                for rs in current_rows:
                    # collect unique children & groups for this row from all nodes in the current frontier
                    children = []
                    grp = []
                    seen_child = set()
                    seen_group = set()

                    for node in rs.frontier:
                        # unique children for this parent frontier; exclude already visited to avoid loops
                        for child in adj_map.get(node, []):
                            if child not in rs.visited and child not in seen_child:
                                seen_child.add(child)
                                children.append(child)
                        # unique groups at this level
                        for g in group_map.get(node, []):
                            if g not in seen_group:
                                seen_group.add(g)
                                grp.append(g)

                    # If nothing to emit at this level, row is terminal
                    if not children and not grp:
                        records.append(rs.out)
                        continue

                    # Explode children one-per-row
                    if children:
                        for child in children:
                            new_rs = deepcopy(rs)
                            chunks_part = f"chunks=[{child}]"
                            groups_part = f"groups=[{', '.join(grp)}]" if grp else ""
                            cell = chunks_part + ("; " if chunks_part and groups_part else "") + groups_part
                            new_rs.out[f"Level_{level_idx}"] = cell
                            new_rs.frontier = [child]
                            new_rs.visited.add(child)
                            next_rows.append(new_rs)
                    else:
                        # groups-only terminal
                        if include_last_groups_only_level and grp:
                            rs.out[f"Level_{level_idx}"] = f"groups=[{', '.join(grp)}]"
                        records.append(rs.out)

                # prepare next iteration
                current_rows = next_rows
                if not current_rows:
                    break
                level_idx += 1
                depth_guard += 1

            # if loop exited by depth guard, flush remaining rows
            if depth_guard >= max_depth:
                for rs in current_rows:
                    records.append(rs.out)

        # Create DataFrame
        out_df = pd.DataFrame(records)

        # Stable sort by all existing parent columns to keep branches contiguous
        sort_cols = [c for c in out_df.columns if c.startswith("Level_")]
        sort_cols = ["root_chunk"] + sorted(sort_cols, key=lambda x: int(x.split("_")[1]))
        sort_cols = [c for c in sort_cols if c in out_df.columns]
        out_df = out_df.sort_values(by=sort_cols, kind="stable").reset_index(drop=True)

        # Ensure ordered columns for output
        level_cols_sorted = [c for c in sort_cols if c.startswith("Level_")]
        out_df = out_df[["root_chunk"] + level_cols_sorted] if level_cols_sorted else out_df[["root_chunk"]]
        return out_df

    # ===============================
    # 6) Write to Excel and merge identical parent cells in ALL levels
    # ===============================
    def merge_identical_cells(ws, col_idx: int, start_row: int, end_row: int):
        """
        Merge consecutive identical cells vertically in a given column from start_row to end_row (inclusive).
        """
        if end_row < start_row:
            return
        col_letter = get_column_letter(col_idx)
        block_start = start_row
        prev_val = ws[f"{col_letter}{start_row}"].value

        for r in range(start_row + 1, end_row + 1):
            cur_val = ws[f"{col_letter}{r}"].value
            if cur_val != prev_val:
                # close previous block if length > 1 and value is not None/empty
                if prev_val not in (None, "", "nan"):
                    if r - 1 > block_start:
                        ws.merge_cells(start_row=block_start, start_column=col_idx, end_row=r - 1, end_column=col_idx)
                        ws[f"{col_letter}{block_start}"].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                # start new block
                block_start = r
                prev_val = cur_val

        # close last block
        if prev_val not in (None, "", "nan"):
            if end_row > block_start:
                ws.merge_cells(start_row=block_start, start_column=col_idx, end_row=end_row, end_column=col_idx)
                ws[f"{col_letter}{block_start}"].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    # Build flow (rows exploded; de-dup per parent)
    flow_df = build_levelwise_flow_rows(
        df,
        include_last_groups_only_level=INCLUDE_LAST_GROUPS_ONLY_LEVEL,
        max_depth=MAX_DEPTH
    )

    # Save first, then merge cells
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        sheet_name = "flow_rows"
        flow_df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Get the underlying worksheet to merge cells
        ws = writer.sheets[sheet_name]

        # Decide which columns to merge (ALL parents)
        cols = list(flow_df.columns)
        if MERGE_PARENT_CELLS_MODE == "root_only":
            cols_to_merge = ["root_chunk"]
        else:
            cols_to_merge = cols  # merge every column that has repeating consecutive values

        # Map column names to Excel column indices (1-based)
        col_to_idx = {name: idx + 1 for idx, name in enumerate(cols)}

        first_data_row = 2  # row 1 is header
        last_data_row = first_data_row + len(flow_df) - 1

        for cname in cols_to_merge:
            cidx = col_to_idx[cname]
            merge_identical_cells(ws, cidx, first_data_row, last_data_row)

    print(f"✅ Generated {len(flow_df)} row(s) across {flow_df['root_chunk'].nunique()} root(s).")
    print(f"   Wrote to: {OUTPUT_XLSX}")
    print(f"   Parent cells merged mode: {MERGE_PARENT_CELLS_MODE}")



