from openpyxl import load_workbook
from copy import copy

input_file = "input.xlsx"
output_file = "output_display_fixed.xlsx"

sheet_name = None      # Keep None to use active sheet, or give sheet name like "Sheet1"
header_row = 1         # Change if your headers are not in row 1


def copy_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def get_merged_range_for_cell(ws, row, col):
    """
    Returns merged range if the cell belongs to a merged range.
    Otherwise returns None.
    """
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return merged_range

    return None


def get_display_value(ws, row, col):
    """
    Returns the visible/displayed value of a cell.
    If the cell is part of a merged range, returns the top-left value.
    """
    merged_range = get_merged_range_for_cell(ws, row, col)

    if merged_range:
        return ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col
        ).value

    return ws.cell(row=row, column=col).value


def get_parent_segments(ws, start_row, end_row, parent_col):
    """
    Splits a row range into segments based on value changes in previous column.

    Example:
    Rows 2-5 in parent column:
        Row 2-3 -> C##.M2
        Row 4-5 -> C##.M3

    Output:
        [(2, 3, 'C##.M2'), (4, 5, 'C##.M3')]
    """
    segments = []

    segment_start = start_row
    previous_value = get_display_value(ws, start_row, parent_col)

    for row in range(start_row + 1, end_row + 1):
        current_value = get_display_value(ws, row, parent_col)

        if current_value != previous_value:
            segments.append((segment_start, row - 1, previous_value))
            segment_start = row
            previous_value = current_value

    segments.append((segment_start, end_row, previous_value))

    return segments


def split_child_merge_by_previous_column(ws, child_range):
    """
    Splits a child merged cell range only if the previous column has
    multiple different parent values within the same row range.
    """
    child_col = child_range.min_col
    parent_col = child_col - 1

    # First column has no previous parent column
    if parent_col < 1:
        return

    # Only handle vertical merged cells in one column
    if child_range.min_col != child_range.max_col:
        return

    start_row = child_range.min_row
    end_row = child_range.max_row

    # Ignore header row merges if any
    if end_row <= header_row:
        return

    child_cell = ws.cell(row=start_row, column=child_col)
    child_value = child_cell.value

    # Nothing to split if child has no value
    if child_value is None:
        return

    parent_segments = get_parent_segments(ws, start_row, end_row, parent_col)

    # Remove blank/None parent-only segments if needed
    # If you want blanks also to be treated as parents, remove this line.
    parent_segments = [
        segment for segment in parent_segments
        if segment[2] is not None
    ]

    # If only one parent value exists inside this child range, no change needed
    if len(parent_segments) <= 1:
        return

    # Save original style
    source_cell = ws.cell(row=start_row, column=child_col)
    original_font = copy(source_cell.font)
    original_fill = copy(source_cell.fill)
    original_border = copy(source_cell.border)
    original_alignment = copy(source_cell.alignment)
    original_number_format = source_cell.number_format
    original_protection = copy(source_cell.protection)

    # Unmerge only this child range, not the full sheet
    ws.unmerge_cells(str(child_range))

    # Clear values in the old child range
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=child_col)
        cell.value = None
        cell.font = copy(original_font)
        cell.fill = copy(original_fill)
        cell.border = copy(original_border)
        cell.alignment = copy(original_alignment)
        cell.number_format = original_number_format
        cell.protection = copy(original_protection)

    # Recreate child merges according to parent segments
    for seg_start, seg_end, parent_value in parent_segments:
        target_cell = ws.cell(row=seg_start, column=child_col)
        target_cell.value = child_value

        target_cell.font = copy(original_font)
        target_cell.fill = copy(original_fill)
        target_cell.border = copy(original_border)
        target_cell.alignment = copy(original_alignment)
        target_cell.number_format = original_number_format
        target_cell.protection = copy(original_protection)

        # Merge only if more than one row
        if seg_start < seg_end:
            ws.merge_cells(
                start_row=seg_start,
                start_column=child_col,
                end_row=seg_end,
                end_column=child_col
            )


def fix_display_merges(ws):
    """
    Generic function.
    Works across all used columns and rows.
    Processes child columns from left to right.
    """
    max_col = ws.max_column

    # Process column by column from left to right
    for child_col in range(2, max_col + 1):

        # Take a fresh copy each time because merge ranges change during processing
        merged_ranges = list(ws.merged_cells.ranges)

        for merged_range in merged_ranges:
            if (
                merged_range.min_col == child_col
                and merged_range.max_col == child_col
            ):
                split_child_merge_by_previous_column(ws, merged_range)


wb = load_workbook(input_file)

if sheet_name:
    ws = wb[sheet_name]
else:
    ws = wb.active

fix_display_merges(ws)

wb.save(output_file)

print(f"Done. Saved file as: {output_file}")
